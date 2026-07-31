"""
automation/sync/sor_sync.py

Utilities for synchronising automation outputs with the canonical
System of Record (SoR) Excel workbook.

Requires the optional ``openpyxl`` dependency:
    pip install openpyxl
"""

from __future__ import annotations

import os
from typing import Any, Dict, List, Optional

CANONICAL_SHEETS = [
    "Roles",
    "Companies",
    "Contacts",
    "Outreach",
    "Interviews",
    "Consulting",
    "Metrics",
    "StatusHistory",
    "FlowErrors",
    "ChangeLog",
]


def _assert_canonical(sheet_name: str) -> None:
    if sheet_name not in CANONICAL_SHEETS:
        raise ValueError(
            f"'{sheet_name}' is not a canonical SoR sheet. "
            f"Allowed: {CANONICAL_SHEETS}"
        )


def read_sheet(sheet_name: str, sor_path: str) -> List[Dict[str, Any]]:
    """Read all rows from a canonical SoR sheet as a list of dicts.

    Args:
        sheet_name: One of the 10 canonical sheet names.
        sor_path: Path to the SoR Excel workbook.

    Returns:
        List of row dicts keyed by column header.

    Raises:
        ValueError: If ``sheet_name`` is not canonical.
        ImportError: If ``openpyxl`` is not installed.
        FileNotFoundError: If ``sor_path`` does not exist.
    """
    _assert_canonical(sheet_name)
    try:
        import openpyxl  # type: ignore
    except ImportError as exc:
        raise ImportError("openpyxl is required for SoR sync. Run: pip install openpyxl") from exc

    if not os.path.exists(sor_path):
        raise FileNotFoundError(f"SoR workbook not found: {sor_path}")

    wb = openpyxl.load_workbook(sor_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h) if h is not None else "" for h in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:]]


def write_rows(
    sheet_name: str,
    rows: List[Dict[str, Any]],
    sor_path: str,
    changed_by: str = "sor_sync",
) -> int:
    """Append rows to a canonical SoR sheet and log each write to ChangeLog.

    Args:
        sheet_name: One of the 10 canonical sheet names.
        rows: List of row dicts. Keys must match the sheet's column headers.
        sor_path: Path to the SoR Excel workbook.
        changed_by: Actor name written to the ChangeLog.

    Returns:
        Number of rows written.

    Raises:
        ValueError: If ``sheet_name`` is not canonical.
        ImportError: If ``openpyxl`` is not installed.
        FileNotFoundError: If ``sor_path`` does not exist.
    """
    _assert_canonical(sheet_name)
    try:
        import openpyxl  # type: ignore
    except ImportError as exc:
        raise ImportError("openpyxl is required for SoR sync. Run: pip install openpyxl") from exc

    if not os.path.exists(sor_path):
        raise FileNotFoundError(f"SoR workbook not found: {sor_path}")

    from automation.core.sor import log_change  # local import to avoid circular deps

    wb = openpyxl.load_workbook(sor_path)
    ws = wb[sheet_name]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    for row in rows:
        ws.append([row.get(h) for h in headers])
        log_change(
            sheet_name=sheet_name,
            field_name="(row append)",
            old_value="",
            new_value=str(row),
            changed_by=changed_by,
        )

    wb.save(sor_path)
    return len(rows)
